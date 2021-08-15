from openpyxl import load_workbook
import numpy as np
# 預設可讀寫，若有需要可以指定write_only和read_only為True
wb = load_workbook('card_list.xlsx')

def get_table(table_name):
	sheet = wb.get_sheet_by_name(table_name)
	return sheet

def set_oda():
	table_list = []
	for column in sheet.columns:
		for cell in column:
			table_list.append(cell.value)
	return table_list

def set_oda_to_tda(oda):
	table = [[0]*sheet.max_row for i in range(sheet.max_column)]
	table = np.array(oda).reshape(sheet.max_column, sheet.max_row)
	return table
		
def set_table():
	table_list = set_oda()
	table = set_oda_to_tda(table_list)
	return table

def search_table(number):
	count = 0
	for name_value in name_table[0][1:]:
		count = count + 1
		if name_value == number:
			show_table(name_table[1][count], min_table, 0)
			show_table('初始', min_table, count)
			show_table('滿等', max_table, count)

def show_table(title, table, value_index):
	text_list = []
	for value in table:
		text_list.append(value[value_index])
	print(title, text_list)

sheet = get_table('Min')
min_table = set_table()

sheet = get_table('Max')
max_table = set_table()

sheet = get_table('Name')
name_table = set_table()

search_table('5')