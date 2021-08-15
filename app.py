from appium import webdriver
import aircv as ac
import time
import os

from openpyxl import load_workbook
import numpy as np

imsrc_name = None
imobj_name = None

wb = load_workbook('card_list.xlsx')
def get_table(table_name):
	sheet = wb.get_sheet_by_name(table_name)
	return sheet

def set_oda():
	table_list = []
	for column in sheet.columns:
		for cell in column:
			table_list.append(cell.value)
	return table_list

def set_oda_to_tda(oda):
	table = [[0]*sheet.max_row for i in range(sheet.max_column)]
	table = np.array(oda).reshape(sheet.max_column, sheet.max_row)
	return table
		
def set_table():
	table_list = set_oda()
	table = set_oda_to_tda(table_list)
	return table

def search_table(number):
	count = 0
	for name_value in name_table[0][1:]:
		count = count + 1
		if name_value == number:
			show_table(name_table[1][count], min_table, 0)
			show_table('初始', min_table, count)
			show_table('滿等', max_table, count)

def show_table(title, table, value_index):
	text_list = []
	for value in table:
		text_list.append(value[value_index])
	print(title, text_list)

sheet = get_table('Min')
min_table = set_table()

sheet = get_table('Max')
max_table = set_table()

sheet = get_table('Name')
name_table = set_table()

img_location = [['confirm', 'confirm_button'], ['ani', 'ani_skip'], ['resoult', 'resoult_button'], ['test', 'test1']]

desired_caps = {}
desired_caps['platformName'] = "Android"         # 声明是ios还是Android系统
desired_caps['platformVersion'] = '7.1.2'        # Android内核版本号，可以在夜神模拟器设置中查看   
desired_caps['deviceName'] = '127.0.0.1:62001'   # 连接的设备名称
desired_caps['noReset'] = True
desired_caps['appPackage'] = 'com.userjoy.sin_mcard'    # apk的包名
desired_caps['appActivity'] = 'com.unity3d.player.UnityPlayerActivity'  # apk的launcherActivity
desired_caps['newCommandTimeout'] = '10000'

driver = webdriver.Remote('http://localhost:4723/wd/hub', desired_caps)          # 建立 session

def set_img_name(img_name):
	img_fullname = img_name + '.png'
	return img_fullname

def get_size():
	phone_screen_size_width = driver.get_window_size()['width']
	phone_screen_size_height = driver.get_window_size()['height']
	return(phone_screen_size_width, phone_screen_size_height)

def screen_shot(folder_path):
	i = 0
	save_file = False
	while save_file != True:
		filename = ('screen_shot' + str(i) + '.png')
		if folder_path == 'log':
			filename = (os.path.join('log', filename))		
		if os.path.isfile(filename):
			i = i + 1
		else:
			driver.get_screenshot_as_file(filename)#截屏手機
			save_file = True

def matchImg(confidence, mode):#imgsrc=原始圖像，imgobj=待查找的圖片
	imsrc = ''
	imobj = ''
	global imsrc_name
	global imobj_name
	if mode == 'UI':
		imsrc = ac.imread(os.path.join('UI_Position', imsrc_name))
		imobj = ac.imread(os.path.join('UI_Position', imobj_name))
	elif mode == 'Check':
		imsrc = ac.imread(os.path.join('Log', imsrc_name))
		imobj = ac.imread(os.path.join('Card', imobj_name))
	else:
		imsrc_name = set_img_name(img_location[3][0])
		imobj_name = set_img_name(img_location[3][1])

	match_result = ac.find_template(imsrc, imobj, confidence)
	if match_result is not None:
		match_result['shape']=(imsrc.shape[1],imsrc.shape[0])#0為高，1為寬

	return match_result

def tab_image():
	confidencevalue = 0.8  # 定義相似度
	position = matchImg(confidencevalue, 'UI')# 用第一步的方法，實際就是find_template()方法

	if position != None:
		x, y = position['result']
		driver.tap([(x, y)])

def swip_right(t):
	screen_size = get_size()
	x1 = int(screen_size[0] * 0.05)
	y1 = int(screen_size[1] * 0.5)
	x2 = int(screen_size[0] * 0.75)
	driver.swipe(x1, y1, x2, y1, t)

while True:
	time_count =0
	mode = input('Enter operator(tap, swip, shot, auto, q): ')
	if mode == 'tap':
		#driver.find_elements_by_id('com.userjoy.sin_mcard:id/view_platform_item_v')[0].click()
		tab_image()
	if mode == 'match':		
		print(matchImg(0.5, ''))
	if mode == 'shot':
		screen_shot('')
	if mode == 'swip':
		swip_right(100)
	if mode == 'auto':
		loop_time = input('How many times do you want to loop: ')
		while time_count < int(loop_time):
			time.sleep(2)
			imsrc_name = set_img_name(img_location[0][0])
			imobj_name = set_img_name(img_location[0][1])
			tab_image()

			time.sleep(3)
			swip_right(100)

			time.sleep(1)
			imsrc_name = set_img_name(img_location[1][0])
			imobj_name = set_img_name(img_location[1][1])
			tab_image()

			time.sleep(1)
			tab_image()

			time.sleep(2)
			screen_shot('log')
			imsrc_name = set_img_name(img_location[2][0])
			imobj_name = set_img_name(img_location[2][1])
			tab_image()
			time_count = time_count + 1
	if mode == 'q':
		print('Exit')
		driver.quit()      # 退出 session
		time.sleep(5)
		break

# def main():
# 	tab_image()

# main()